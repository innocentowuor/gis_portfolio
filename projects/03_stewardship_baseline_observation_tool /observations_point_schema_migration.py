"""
Mountain Goat observations schema migration and export.

Purpose
-------
This script demonstrates how several legacy wildlife-observation layers can be
standardized into one submission-ready schema. It shows the main workflow from
an internal project without including real service URLs, item IDs, network
paths, credentials, staff information, place names, observation records, or
the complete production schema.

Workflow
--------
1. Download three legacy point layers from ArcGIS Online.
2. Project each layer to a shared coordinate system.
3. Create an empty feature class with a simplified target schema.
4. Map source fields into that schema and add standard values.
5. Normalize dates, seasons, activity codes, sign codes, ratings, and counts.
6. Add generalized area attributes with spatial joins.
7. Export records into annual, area-based Excel submission files.

This is a simplified portfolio sample, not a drop-in production script.
Replace the placeholder services, reference layers, field mappings, and Excel
template before running it with authorized data.

Required software
-----------------
- ArcGIS Pro Python environment (arcpy and arcgis)
- pandas
- openpyxl

Credentials and paths are read from environment variables. No custom classes
are used; the workflow is organized with functions, dictionaries, and loops.
"""

import logging
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

import arcpy
import pandas as pd
from arcgis import GIS
from arcgis.features import FeatureLayer
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

PORTAL_URL = os.getenv("GIS_PORTAL_URL", "https://www.arcgis.com")
GIS_USERNAME = os.getenv("GIS_USERNAME")
GIS_PASSWORD = os.getenv("GIS_PASSWORD")

WORKSPACE = Path(os.getenv("WILDLIFE_WORKSPACE", r"C:\GIS\wildlife_schema_demo"))
SCRATCH_GDB = WORKSPACE / "scratch.gdb"
OUTPUT_GDB = WORKSPACE / "output.gdb"
EXPORT_FOLDER = WORKSPACE / "submission_exports"
TEMPLATE_FILE = WORKSPACE / "templates" / "wildlife_submission_template.xlsx"

# Placeholder URLs intentionally do not point to real services.
SOURCE_SERVICES = {
    "winter_survey": os.getenv(
        "WINTER_SURVEY_URL",
        "https://example.com/arcgis/rest/services/winter_survey/FeatureServer/0",
    ),
    "summer_survey": os.getenv(
        "SUMMER_SURVEY_URL",
        "https://example.com/arcgis/rest/services/summer_survey/FeatureServer/0",
    ),
    "incidental": os.getenv(
        "INCIDENTAL_URL",
        "https://example.com/arcgis/rest/services/incidental/FeatureServer/0",
    ),
}

# These authorized layers are supplied by the person running the script.
REFERENCE_LAYERS = {
    "management_area": os.getenv("MANAGEMENT_AREA_LAYER", ""),
    "reporting_region": os.getenv("REPORTING_REGION_LAYER", ""),
    "habitat_unit": os.getenv("HABITAT_UNIT_LAYER", ""),
}

TARGET_SPATIAL_REFERENCE = arcpy.SpatialReference(3005)
TARGET_FEATURE_CLASS_NAME = "standardized_wildlife_observations"

# The target is deliberately representative rather than a copy of the internal
# production schema.
TARGET_FIELDS = {
    "study_area": ("TEXT", 100),
    "survey_block": ("TEXT", 100),
    "observer": ("TEXT", 100),
    "species_code": ("TEXT", 20),
    "observation_count": ("LONG", None),
    "adult_males": ("LONG", None),
    "adult_females": ("LONG", None),
    "adults_unclassified": ("LONG", None),
    "juveniles_unclassified": ("LONG", None),
    "yearlings_unclassified": ("LONG", None),
    "life_stage_unclassified": ("LONG", None),
    "activity_code": ("TEXT", 20),
    "sign_code": ("TEXT", 20),
    "season": ("TEXT", 20),
    "observation_type": ("TEXT", 20),
    "observation_date": ("DATE", None),
    "observation_date_text": ("TEXT", 20),
    "snow_cover_rating": ("TEXT", 30),
    "canopy_cover_rating": ("TEXT", 30),
    "terrain_rating": ("TEXT", 30),
    "comments": ("TEXT", 500),
    "management_area": ("TEXT", 100),
    "reporting_region": ("TEXT", 100),
    "habitat_unit": ("TEXT", 100),
    "source_name": ("TEXT", 100),
}

# Synthetic source names show how differently structured legacy layers can be
# brought into one schema.
SOURCE_FIELD_MAPPINGS = {
    "winter_survey": {
        "survey_date": "observation_date",
        "adult_count": "adults_unclassified",
        "juvenile_count": "juveniles_unclassified",
        "unknown_count": "life_stage_unclassified",
        "total_count": "observation_count",
        "observed_activity": "activity_code",
        "tracks_present": "sign_code",
        "snow_rating": "snow_cover_rating",
        "canopy_rating": "canopy_cover_rating",
        "terrain_class": "terrain_rating",
        "notes": "comments",
        "survey_area": "study_area",
        "survey_block": "survey_block",
        "source_dataset": "source_name",
    },
    "summer_survey": {
        "observation_date": "observation_date",
        "adults": "adults_unclassified",
        "juveniles": "juveniles_unclassified",
        "unclassified": "life_stage_unclassified",
        "total": "observation_count",
        "tracks": "sign_code",
        "region": "study_area",
        "source": "source_name",
    },
    "incidental": {
        "date_seen": "observation_date",
        "adult_count": "adults_unclassified",
        "juvenile_count": "juveniles_unclassified",
        "yearling_count": "yearlings_unclassified",
        "unknown_count": "life_stage_unclassified",
        "total_count": "observation_count",
        "activity": "activity_code",
        "tracks": "sign_code",
        "notes": "comments",
        "related_survey": "source_name",
    },
}

SOURCE_DEFAULTS = {
    "winter_survey": {
        "species_code": "TARGET_SPECIES",
        "observation_type": "Survey",
        "season": "Winter",
    },
    "summer_survey": {
        "species_code": "TARGET_SPECIES",
        "observation_type": "Survey",
        "season": "Non-Winter",
    },
    "incidental": {
        "species_code": "TARGET_SPECIES",
        "observation_type": "Incidental",
        "season": "Unknown",
    },
}

ACTIVITY_CODES = {
    "standing": "ST",
    "moving": "MV",
    "bedded": "BD",
    "running": "RN",
    "lying down": "BD",
}

COUNT_FIELDS = [
    "adult_males",
    "adult_females",
    "adults_unclassified",
    "juveniles_unclassified",
    "yearlings_unclassified",
    "life_stage_unclassified",
]

SURVEY_EXPORT_MAPPING = {
    "study_area": "Study Area",
    "survey_block": "Survey Block",
    "observation_date": "Detection Date",
    "species_code": "Species Code",
    "observation_count": "Count",
    "adult_males": "Adult Males",
    "adult_females": "Adult Females",
    "adults_unclassified": "Adults - Unclassified",
    "juveniles_unclassified": "Juveniles - Unclassified",
    "yearlings_unclassified": "Yearlings - Unclassified",
    "life_stage_unclassified": "Life Stage - Unclassified",
    "activity_code": "Activity Code",
    "sign_code": "Sign Code",
    "season": "Season",
    "management_area": "Management Area",
    "habitat_unit": "Habitat Unit",
    "snow_cover_rating": "Snow Cover Rating",
    "canopy_cover_rating": "Canopy Cover Rating",
    "terrain_rating": "Terrain Rating",
}

INCIDENTAL_EXPORT_MAPPING = {
    "study_area": "Study Area",
    "survey_block": "Survey Block",
    "observation_date": "Date & Time",
    "species_code": "Species Code",
    "observation_count": "Count",
    "activity_code": "Activity Code",
    "adult_males": "Adult Males",
    "adult_females": "Adult Females",
    "adults_unclassified": "Adults - Unclassified",
    "juveniles_unclassified": "Juveniles - Unclassified",
    "yearlings_unclassified": "Yearlings - Unclassified",
    "life_stage_unclassified": "Life Stage - Unclassified",
    "comments": "Comments",
    "management_area": "Management Area",
    "habitat_unit": "Habitat Unit",
}


# ---------------------------------------------------------------------------
# Setup and validation
# ---------------------------------------------------------------------------

def configure_logging():
    """Send readable progress messages to the console."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )


def validate_configuration():
    """Stop early when required local settings are missing."""
    missing = []

    if not GIS_USERNAME or not GIS_PASSWORD:
        missing.append("GIS_USERNAME and GIS_PASSWORD")

    if not TEMPLATE_FILE.exists():
        missing.append(f"submission template: {TEMPLATE_FILE}")

    for name, path in REFERENCE_LAYERS.items():
        if not path:
            missing.append(f"{name.upper()}_LAYER")

    if missing:
        raise ValueError(
            "Configure the following before running this example:\n- "
            + "\n- ".join(missing)
        )


def create_file_geodatabase(gdb_path):
    """Create a file geodatabase when it does not already exist."""
    gdb_path = Path(gdb_path)
    gdb_path.parent.mkdir(parents=True, exist_ok=True)

    if not arcpy.Exists(str(gdb_path)):
        arcpy.management.CreateFileGDB(str(gdb_path.parent), gdb_path.name)


def prepare_workspace():
    """Create output folders and geodatabases."""
    EXPORT_FOLDER.mkdir(parents=True, exist_ok=True)
    create_file_geodatabase(SCRATCH_GDB)
    create_file_geodatabase(OUTPUT_GDB)
    arcpy.env.workspace = str(SCRATCH_GDB)
    arcpy.env.overwriteOutput = True


def connect_to_portal():
    """Connect using credentials stored outside the script."""
    logging.info("Connecting to the configured GIS portal")
    return GIS(PORTAL_URL, GIS_USERNAME, GIS_PASSWORD)


# ---------------------------------------------------------------------------
# Extract and schema creation
# ---------------------------------------------------------------------------

def download_and_project_layer(gis, service_url, output_name):
    """Download one feature layer and project it to the target coordinate system."""
    raw_name = f"{output_name}_raw"
    raw_path = str(SCRATCH_GDB / raw_name)
    projected_path = str(SCRATCH_GDB / output_name)

    for path in [raw_path, projected_path]:
        if arcpy.Exists(path):
            arcpy.management.Delete(path)

    logging.info("Downloading source layer: %s", output_name)
    feature_set = FeatureLayer(service_url, gis=gis).query(
        where="1=1",
        out_fields="*",
        return_geometry=True,
    )
    feature_set.save(str(SCRATCH_GDB), raw_name)

    arcpy.management.Project(
        in_dataset=raw_path,
        out_dataset=projected_path,
        out_coor_system=TARGET_SPATIAL_REFERENCE,
    )
    arcpy.management.Delete(raw_path)
    return projected_path


def download_source_layers(gis):
    """Download all configured legacy datasets."""
    downloaded = {}
    for source_name, service_url in SOURCE_SERVICES.items():
        downloaded[source_name] = download_and_project_layer(
            gis,
            service_url,
            f"legacy_{source_name}",
        )
    return downloaded


def create_target_feature_class():
    """Create an empty point feature class with the simplified target schema."""
    target_path = str(SCRATCH_GDB / TARGET_FEATURE_CLASS_NAME)

    if arcpy.Exists(target_path):
        arcpy.management.Delete(target_path)

    arcpy.management.CreateFeatureclass(
        out_path=str(SCRATCH_GDB),
        out_name=TARGET_FEATURE_CLASS_NAME,
        geometry_type="POINT",
        spatial_reference=TARGET_SPATIAL_REFERENCE,
    )

    for field_name, (field_type, field_length) in TARGET_FIELDS.items():
        add_field_options = {
            "in_table": target_path,
            "field_name": field_name,
            "field_type": field_type,
        }
        if field_length:
            add_field_options["field_length"] = field_length

        arcpy.management.AddField(
            **add_field_options,
        )

    return target_path


# ---------------------------------------------------------------------------
# Transform and load
# ---------------------------------------------------------------------------

def append_mapped_records(source_path, target_path, mapping, defaults):
    """Copy one legacy schema into the common target schema."""
    source_fields = list(mapping.keys())
    mapped_target_fields = list(mapping.values())
    default_fields = list(defaults.keys())
    insert_fields = mapped_target_fields + default_fields + ["SHAPE@"]

    inserted_count = 0
    with arcpy.da.SearchCursor(
        source_path,
        source_fields + ["SHAPE@"],
    ) as search_cursor, arcpy.da.InsertCursor(
        target_path,
        insert_fields,
    ) as insert_cursor:
        for source_row in search_cursor:
            mapped_values = list(source_row[:-1])
            default_values = [defaults[field] for field in default_fields]
            insert_cursor.insertRow(mapped_values + default_values + [source_row[-1]])
            inserted_count += 1

    logging.info("Appended %s records from %s", inserted_count, Path(source_path).name)


def append_all_sources(source_layers, target_path):
    """Apply the correct mapping and defaults to every source layer."""
    for source_name, source_path in source_layers.items():
        append_mapped_records(
            source_path,
            target_path,
            SOURCE_FIELD_MAPPINGS[source_name],
            SOURCE_DEFAULTS[source_name],
        )


def season_from_date(observation_date):
    """Classify November through April as winter."""
    if not observation_date:
        return "Unknown"
    if observation_date.month in [11, 12, 1, 2, 3, 4]:
        return "Winter"
    return "Non-Winter"


def normalize_activity(value):
    """Convert descriptive activity values to short submission codes."""
    if not value:
        return None
    cleaned = str(value).strip()
    return ACTIVITY_CODES.get(cleaned.lower(), cleaned)


def normalize_sign(value):
    """Convert yes/no track indicators to a standard sign code."""
    if value is None:
        return None
    cleaned = str(value).strip().lower()
    if cleaned in {"yes", "y", "true", "1"}:
        return "TR"
    if cleaned in {"no", "n", "false", "0", ""}:
        return None
    return str(value).strip()


def text_before_bracket(value):
    """Keep the useful portion of a rating such as 'High (76-100%)'."""
    if not value:
        return None
    return re.split(r"[\(\[]", str(value), maxsplit=1)[0].strip() or None


def standardize_date(value):
    """Use a consistent time so date-only observations survive UTC conversion."""
    if not value:
        return None
    return datetime.combine(value.date(), datetime.min.time()) + timedelta(hours=12)


def normalize_target_values(target_path):
    """Clean coded values, recalculate totals, and synchronize date fields."""
    cursor_fields = [
        "observation_date",
        "observation_date_text",
        "season",
        "activity_code",
        "sign_code",
        "snow_cover_rating",
        "canopy_cover_rating",
    ] + COUNT_FIELDS + ["observation_count"]

    with arcpy.da.UpdateCursor(target_path, cursor_fields) as cursor:
        for row in cursor:
            observation_date = standardize_date(row[0])
            row[0] = observation_date
            row[1] = observation_date.strftime("%Y-%m-%d") if observation_date else None
            row[2] = season_from_date(observation_date)
            row[3] = normalize_activity(row[3])
            row[4] = normalize_sign(row[4])
            row[5] = text_before_bracket(row[5])
            row[6] = text_before_bracket(row[6])

            count_start = 7
            count_end = count_start + len(COUNT_FIELDS)
            row[-1] = sum(value or 0 for value in row[count_start:count_end])
            cursor.updateRow(row)


# ---------------------------------------------------------------------------
# Spatial enrichment
# ---------------------------------------------------------------------------

def spatially_add_attribute(
    input_features,
    join_features,
    join_field,
    target_field,
    output_name,
    match_option="INTERSECT",
    search_radius=None,
):
    """Spatially join one generalized reference attribute to the observations."""
    output_path = str(SCRATCH_GDB / output_name)
    if arcpy.Exists(output_path):
        arcpy.management.Delete(output_path)

    arcpy.analysis.SpatialJoin(
        target_features=input_features,
        join_features=join_features,
        out_feature_class=output_path,
        join_operation="JOIN_ONE_TO_ONE",
        join_type="KEEP_ALL",
        match_option=match_option,
        search_radius=search_radius,
    )

    available_fields = {field.name for field in arcpy.ListFields(output_path)}
    if join_field not in available_fields:
        raise ValueError(f"Expected joined field is missing: {join_field}")

    with arcpy.da.UpdateCursor(
        output_path,
        [join_field, target_field],
    ) as cursor:
        for row in cursor:
            if row[0]:
                row[1] = row[0]
                cursor.updateRow(row)

    return output_path


def enrich_observations(target_path):
    """Add three area attributes from authorized reference layers."""
    enriched_path = spatially_add_attribute(
        target_path,
        REFERENCE_LAYERS["management_area"],
        "AREA_NAME",
        "management_area",
        "observations_with_management_area",
    )
    enriched_path = spatially_add_attribute(
        enriched_path,
        REFERENCE_LAYERS["reporting_region"],
        "REGION_NAME",
        "reporting_region",
        "observations_with_reporting_region",
    )
    enriched_path = spatially_add_attribute(
        enriched_path,
        REFERENCE_LAYERS["habitat_unit"],
        "UNIT_NAME",
        "habitat_unit",
        "observations_with_habitat_unit",
        match_option="WITHIN_A_DISTANCE",
        search_radius="100 Meters",
    )
    return enriched_path


def export_final_feature_class(enriched_path):
    """Save the standardized feature class in the output geodatabase."""
    final_path = str(OUTPUT_GDB / TARGET_FEATURE_CLASS_NAME)
    if arcpy.Exists(final_path):
        arcpy.management.Delete(final_path)
    arcpy.conversion.ExportFeatures(enriched_path, final_path)
    return final_path


# ---------------------------------------------------------------------------
# Excel submission exports
# ---------------------------------------------------------------------------

def feature_class_to_dataframe(feature_class):
    """Read non-geometry attributes into a DataFrame."""
    fields = [
        field.name
        for field in arcpy.ListFields(feature_class)
        if field.type not in {"OID", "Geometry", "Blob", "Raster"}
    ]
    rows = list(arcpy.da.SearchCursor(feature_class, fields))
    return pd.DataFrame(rows, columns=fields)


def prepare_export_table(data, field_mapping, template_fields):
    """Rename and order values to match a submission-template sheet."""
    return (
        data.rename(columns=field_mapping)
        .reindex(columns=template_fields, fill_value="")
        .fillna("")
    )


def clear_sheet_data(worksheet):
    """Clear template sample rows while preserving the header and formatting."""
    if worksheet.max_row < 2:
        return
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.value = None


def write_dataframe(worksheet, dataframe):
    """Write a DataFrame beneath an existing template header."""
    for row_number, values in enumerate(
        dataframe_to_rows(dataframe, index=False, header=False),
        start=2,
    ):
        for column_number, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column_number, value=value)


def safe_filename_part(value):
    """Remove characters that are unsafe in Windows filenames."""
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", str(value)).strip()
    return cleaned or "Unassigned"


def export_submission_workbooks(feature_class):
    """Create one workbook per observation year and generalized study area."""
    observations = feature_class_to_dataframe(feature_class)
    observations["observation_date"] = pd.to_datetime(
        observations["observation_date"],
        errors="coerce",
    )
    observations["year"] = observations["observation_date"].dt.year
    observations = observations.dropna(subset=["year", "study_area"])

    template = load_workbook(TEMPLATE_FILE)
    survey_sheet_name = "Survey Observations"
    incidental_sheet_name = "Incidental Observations"

    survey_fields = [
        cell.value
        for cell in template[survey_sheet_name][1]
        if cell.value
    ]
    incidental_fields = [
        cell.value
        for cell in template[incidental_sheet_name][1]
        if cell.value
    ]

    for (year, study_area), group in observations.groupby(["year", "study_area"]):
        survey_rows = group[group["observation_type"] == "Survey"]
        incidental_rows = group[group["observation_type"] == "Incidental"]

        survey_export = prepare_export_table(
            survey_rows,
            SURVEY_EXPORT_MAPPING,
            survey_fields,
        )
        incidental_export = prepare_export_table(
            incidental_rows,
            INCIDENTAL_EXPORT_MAPPING,
            incidental_fields,
        )

        workbook = load_workbook(TEMPLATE_FILE)
        clear_sheet_data(workbook[survey_sheet_name])
        clear_sheet_data(workbook[incidental_sheet_name])
        write_dataframe(workbook[survey_sheet_name], survey_export)
        write_dataframe(workbook[incidental_sheet_name], incidental_export)

        filename = (
            f"{int(year)}_wildlife_observations_"
            f"{safe_filename_part(study_area)}.xlsx"
        )
        output_path = EXPORT_FOLDER / filename
        workbook.save(output_path)
        logging.info("Created submission workbook: %s", output_path.name)


# ---------------------------------------------------------------------------
# Main workflow
# ---------------------------------------------------------------------------

def main():
    """Run the schema migration and submission export workflow."""
    configure_logging()
    validate_configuration()
    prepare_workspace()

    gis = connect_to_portal()
    source_layers = download_source_layers(gis)

    target_path = create_target_feature_class()
    append_all_sources(source_layers, target_path)
    normalize_target_values(target_path)

    enriched_path = enrich_observations(target_path)
    final_path = export_final_feature_class(enriched_path)
    export_submission_workbooks(final_path)

    logging.info("Workflow complete")
    logging.info("Standardized feature class: %s", final_path)
    logging.info("Submission exports: %s", EXPORT_FOLDER)


if __name__ == "__main__":
    main()

