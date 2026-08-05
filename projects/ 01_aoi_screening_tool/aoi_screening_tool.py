"""
Area of Interest Screening Tool — public portfolio example
============================================================

Purpose
-------
This workflow processes polygon submissions from an ArcGIS feature service, compares each area
of interest (AOI) with reference datasets, and creates an Excel screening report
with summary statistics and a map.

Workflow
--------
1. Connect to an ArcGIS portal using environment variables.
2. Find submissions that are ready to process.
3. Obtain each AOI from its feature geometry or an uploaded zipped shapefile.
4. Project the AOI into a common coordinate system.
5. Intersect the AOI with configured point and polygon reference layers.
6. Summarize feature counts, overlap area, and percentage of the AOI.
7. Export a map from an ArcGIS Pro layout and add it to an Excel report.
8. Update the submission status.

Public-safe design
------------------
This is a simplified portfolio version of a production workflow. Organization
names, species names, service item IDs, URLs, paths, field schemas, email
addresses, business rules, and proprietary helper libraries have been removed
or replaced with generic examples. It contains no production data or secrets.

Requirements
------------
- ArcGIS Pro Python environment (``arcpy`` and ``arcgis``)
- pandas and openpyxl
- An ArcGIS Pro project containing a map named ``Screening Map`` and a layout
  named ``Screening Layout``

Configuration
-------------
Set ``SCREENING_PORTAL_URL``, ``SCREENING_USERNAME``, and
``SCREENING_PASSWORD`` as environment variables. Replace the placeholder
service item ID and local reference-layer paths below with resources you are
authorized to use.

This example uses regular functions, dictionaries, and a clear ``main()``
workflow so each step can be followed without object-oriented programming.
It will not run unchanged because all external resources are placeholders.
"""

from __future__ import annotations

import logging
import os
import re
import tempfile
import time
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZipFile

import arcpy
import openpyxl
import pandas as pd
from arcgis.features import FeatureLayer
from arcgis.gis import GIS
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


LOGGER = logging.getLogger("environmental_screening")


def load_settings() -> dict[str, Any]:
    """Load credentials and file locations into a simple settings dictionary."""
    required = {
        "SCREENING_PORTAL_URL": os.getenv("SCREENING_PORTAL_URL"),
        "SCREENING_USERNAME": os.getenv("SCREENING_USERNAME"),
        "SCREENING_PASSWORD": os.getenv("SCREENING_PASSWORD"),
    }
    missing = [name for name, value in required.items() if not value]
    if missing:
        raise RuntimeError(
            "Missing required environment variables: " + ", ".join(missing)
        )

    project_root = Path(__file__).resolve().parent
    return {
        "portal_url": required["SCREENING_PORTAL_URL"],
        "username": required["SCREENING_USERNAME"],
        "password": required["SCREENING_PASSWORD"],
        "submission_item_id": "REPLACE_WITH_AUTHORIZED_ITEM_ID",
        "submission_layer_index": 0,
        "workspace": project_root / "workspace" / "screening.gdb",
        "project_template": project_root / "templates" / "screening_template.aprx",
        "output_directory": project_root / "output",
        "target_wkid": 3005,
    }


# Synthetic layer names and schemas are used intentionally.
# Replace these paths only with public, synthetic, or authorized data.
REFERENCE_LAYERS = [
    {
        "display_name": "Known Sites",
        "path": r"C:\public_demo\data\reference.gdb\known_sites",
        "geometry_type": "point",
        "report_fields": ["site_id", "category", "survey_year"],
    },
    {
        "display_name": "Environmental Observations",
        "path": r"C:\public_demo\data\reference.gdb\observations",
        "geometry_type": "point",
        "report_fields": [
            "observation_id",
            "observation_type",
            "observation_date",
        ],
    },
    {
        "display_name": "Management Areas",
        "path": r"C:\public_demo\data\reference.gdb\management_areas",
        "geometry_type": "polygon",
        "report_fields": ["area_id", "management_class"],
    },
    {
        "display_name": "Habitat Suitability",
        "path": r"C:\public_demo\data\reference.gdb\habitat_suitability",
        "geometry_type": "polygon",
        "report_fields": ["habitat_class"],
    },
]


def configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )


def configure_arcpy(workspace: Path) -> None:
    """Configure ArcPy and ensure the output file geodatabase exists."""
    workspace.parent.mkdir(parents=True, exist_ok=True)
    if not arcpy.Exists(str(workspace)):
        arcpy.management.CreateFileGDB(str(workspace.parent), workspace.name)

    arcpy.env.workspace = str(workspace)
    arcpy.env.overwriteOutput = True
    arcpy.env.parallelProcessingFactor = "100%"


def connect_to_portal(settings: dict[str, Any]) -> GIS:
    """Create an authenticated portal connection."""
    started = time.perf_counter()
    gis = GIS(
        settings["portal_url"],
        settings["username"],
        settings["password"],
    )
    LOGGER.info("Connected to the portal in %.1f seconds", time.perf_counter() - started)
    return gis


def get_submission_layer(
    gis: GIS,
    settings: dict[str, Any],
) -> FeatureLayer:
    """Return the configured layer from the submission feature-service item."""
    item = gis.content.get(settings["submission_item_id"])
    if item is None:
        raise LookupError("The configured submission item was not found.")
    try:
        return item.layers[settings["submission_layer_index"]]
    except IndexError as exc:
        raise LookupError("The configured submission layer does not exist.") from exc


def get_pending_submission_ids(layer: FeatureLayer) -> list[int]:
    """Return IDs for records waiting to be processed."""
    result = layer.query(
        where="processing_status IS NULL OR processing_status IN ('Ready', 'Failed')",
        out_fields="objectid",
        return_geometry=False,
    )
    return [int(feature.attributes["objectid"]) for feature in result.features]


def update_status(
    layer: FeatureLayer,
    object_id: int,
    status: str,
    message: str | None = None,
) -> None:
    """Update workflow fields on one submission."""
    result = layer.query(where=f"objectid = {object_id}", out_fields="*")
    if not result.features:
        raise LookupError(f"Submission {object_id} was not found.")

    feature = result.features[0]
    feature.attributes["processing_status"] = status
    feature.attributes["processing_message"] = message
    response = layer.edit_features(updates=[feature])

    update_result = response.get("updateResults", [{}])[0]
    if not update_result.get("success"):
        raise RuntimeError(f"Could not update submission {object_id}.")


def _safe_extract_zip(zip_path: Path, destination: Path) -> None:
    """Extract a ZIP file while blocking path-traversal entries."""
    destination_root = destination.resolve()
    with ZipFile(zip_path) as archive:
        for member in archive.infolist():
            member_path = (destination / member.filename).resolve()
            if destination_root not in member_path.parents and member_path != destination_root:
                raise ValueError(f"Unsafe ZIP entry: {member.filename}")
        archive.extractall(destination)


def obtain_aoi(
    layer: FeatureLayer,
    object_id: int,
    workspace: Path,
    target_wkid: int,
) -> str:
    """
    Save the submitted AOI to the workspace.

    The example accepts either feature geometry drawn in the form or the first
    zipped shapefile attachment. Production applications should add explicit
    file-size, file-count, geometry, and coordinate-system validation.
    """
    raw_name = f"aoi_{object_id}_source"
    projected_name = f"aoi_{object_id}"
    raw_path = str(workspace / raw_name)
    projected_path = str(workspace / projected_name)

    attachments = layer.attachments.get_list(oid=object_id)
    if attachments:
        with tempfile.TemporaryDirectory(prefix="screening_aoi_") as temp_dir:
            temp_path = Path(temp_dir)
            attachment = attachments[0]
            downloaded = layer.attachments.download(
                oid=object_id,
                attachment_id=attachment["id"],
                save_path=str(temp_path),
            )
            # The ArcGIS API may return either one path or a one-item path list.
            downloaded_path = downloaded[0] if isinstance(downloaded, list) else downloaded
            zip_path = Path(downloaded_path)
            _safe_extract_zip(zip_path, temp_path)
            shapefiles = list(temp_path.rglob("*.shp"))
            if len(shapefiles) != 1:
                raise ValueError("The attachment must contain exactly one shapefile.")
            arcpy.conversion.FeatureClassToFeatureClass(
                str(shapefiles[0]), str(workspace), raw_name
            )
    else:
        feature_set = layer.query(
            where=f"objectid = {object_id}",
            out_fields="objectid",
            return_geometry=True,
        )
        if not feature_set.features:
            raise ValueError("The submission does not contain an AOI geometry.")
        feature_set.save(str(workspace), raw_name)

    description = arcpy.Describe(raw_path)
    if description.shapeType != "Polygon":
        raise ValueError("The submitted AOI must be a polygon.")

    if arcpy.Exists(projected_path):
        arcpy.management.Delete(projected_path)
    arcpy.management.Project(
        raw_path,
        projected_path,
        arcpy.SpatialReference(target_wkid),
    )
    return projected_path


def count_features(feature_class: str) -> int:
    return int(arcpy.management.GetCount(feature_class)[0])


def calculate_area_hectares(feature_class: str) -> float:
    """Sum geodesic feature area and convert square metres to hectares."""
    total_square_metres = 0.0
    with arcpy.da.SearchCursor(feature_class, ["SHAPE@"]) as cursor:
        for (geometry,) in cursor:
            total_square_metres += geometry.getArea("GEODESIC", "SQUAREMETERS")
    return total_square_metres / 10_000


def intersect_layer(
    aoi: str,
    layer: dict[str, Any],
    workspace: Path,
) -> str | None:
    """Intersect one reference layer with the AOI and return its output path."""
    safe_name = re.sub(
        r"[^A-Za-z0-9_]+",
        "_",
        layer["display_name"],
    ).strip("_")
    output = str(workspace / f"intersect_{safe_name}")

    if arcpy.Exists(output):
        arcpy.management.Delete(output)
    arcpy.analysis.Intersect(
        [aoi, layer["path"]],
        output,
        "ALL",
        None,
        "INPUT",
    )

    return output if count_features(output) else None


def existing_fields(feature_class: str) -> set[str]:
    return {field.name for field in arcpy.ListFields(feature_class)}


def read_detail_rows(
    feature_class: str | None,
    requested_fields: Iterable[str],
) -> pd.DataFrame:
    """Read only configured fields that are present in the intersect output."""
    requested = list(requested_fields)
    if feature_class is None:
        return pd.DataFrame(columns=requested)

    available = existing_fields(feature_class)
    fields = [field for field in requested if field in available]
    if not fields:
        return pd.DataFrame()

    rows = list(arcpy.da.SearchCursor(feature_class, fields))
    return pd.DataFrame(rows, columns=fields)


def analyze_submission(
    aoi: str,
    reference_layers: Iterable[dict[str, Any]],
    workspace: Path,
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    """Run all intersections and return summary and detail tables."""
    aoi_area = calculate_area_hectares(aoi)
    if aoi_area <= 0:
        raise ValueError("The submitted AOI has no measurable area.")

    summaries: list[dict[str, Any]] = []
    details: dict[str, pd.DataFrame] = {}

    for layer in reference_layers:
        LOGGER.info("Analyzing %s", layer["display_name"])
        intersection = intersect_layer(aoi, layer, workspace)
        feature_count = count_features(intersection) if intersection else 0

        overlap_area = 0.0
        overlap_percent = 0.0
        if intersection and layer["geometry_type"] == "polygon":
            overlap_area = calculate_area_hectares(intersection)
            overlap_percent = min((overlap_area / aoi_area) * 100, 100.0)

        summaries.append(
            {
                "Reference layer": layer["display_name"],
                "Feature count": feature_count,
                "Overlap area (ha)": round(overlap_area, 2),
                "AOI overlap (%)": round(overlap_percent, 2),
            }
        )
        details[layer["display_name"]] = read_detail_rows(
            intersection,
            layer["report_fields"],
        )

    return pd.DataFrame(summaries), details


def safe_sheet_name(name: str, used_names: set[str]) -> str:
    """Create a valid, unique Excel worksheet name."""
    base = re.sub(r"[\\/*?:\[\]]", "", name).strip()[:31] or "Sheet"
    candidate = base
    counter = 2
    while candidate.casefold() in used_names:
        suffix = f" ({counter})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate.casefold())
    return candidate


def export_map(
    project_template: Path,
    aoi: str,
    output_image: Path,
) -> None:
    """Add the AOI to a template map, zoom to it, and export the layout."""
    project = arcpy.mp.ArcGISProject(str(project_template))
    map_object = project.listMaps("Screening Map")[0]
    layout = project.listLayouts("Screening Layout")[0]
    map_frame = layout.listElements("MAPFRAME_ELEMENT")[0]

    aoi_layer = map_object.addDataFromPath(aoi)
    symbol = aoi_layer.symbology
    if hasattr(symbol.renderer, "symbol"):
        symbol.renderer.symbol.color = {"RGB": [255, 255, 255, 0]}
        symbol.renderer.symbol.outlineColor = {"RGB": [220, 40, 40, 100]}
        symbol.renderer.symbol.outlineSize = 2
        aoi_layer.symbology = symbol

    extent = map_frame.getLayerExtent(aoi_layer, False, True)
    map_frame.camera.setExtent(extent)
    layout.exportToPNG(str(output_image), resolution=200)
    del project


def _style_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in worksheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max((len(value) for value in values), default=8) + 2, 45)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def create_excel_report(
    summary: pd.DataFrame,
    details: dict[str, pd.DataFrame],
    map_image: Path,
    output_file: Path,
) -> None:
    """Create and format a workbook containing a map, summary, and details."""
    used_names = {"map", "summary"}
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        pd.DataFrame({"Screening map": ["See image below"]}).to_excel(
            writer, sheet_name="Map", index=False
        )
        summary.to_excel(writer, sheet_name="Summary", index=False)

        for display_name, table in details.items():
            sheet_name = safe_sheet_name(display_name, used_names)
            if table.empty:
                table = pd.DataFrame({"Result": ["No intersecting features"]})
            table.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = openpyxl.load_workbook(output_file)
    for worksheet in workbook.worksheets:
        _style_worksheet(worksheet)

    image = Image(str(map_image))
    image.anchor = "A3"
    workbook["Map"].add_image(image)
    workbook.save(output_file)


def process_submission(
    submission_layer: FeatureLayer,
    object_id: int,
    settings: dict[str, Any],
) -> Path:
    """Run the complete workflow for one submission."""
    update_status(submission_layer, object_id, "Processing")
    try:
        aoi = obtain_aoi(
            submission_layer,
            object_id,
            settings["workspace"],
            settings["target_wkid"],
        )
        summary, details = analyze_submission(
            aoi,
            REFERENCE_LAYERS,
            settings["workspace"],
        )

        output_directory = settings["output_directory"]
        output_directory.mkdir(parents=True, exist_ok=True)
        map_image = output_directory / f"screening_map_{object_id}.png"
        report = output_directory / f"screening_report_{object_id}.xlsx"

        export_map(settings["project_template"], aoi, map_image)
        create_excel_report(summary, details, map_image, report)
        update_status(submission_layer, object_id, "Complete")
        return report
    except Exception as exc:
        LOGGER.exception("Submission %s failed", object_id)
        update_status(
            submission_layer,
            object_id,
            "Failed",
            message=str(exc)[:250],
        )
        raise


def main() -> None:
    """Process every pending submission independently."""
    configure_logging()
    settings = load_settings()
    configure_arcpy(settings["workspace"])

    gis = connect_to_portal(settings)
    submission_layer = get_submission_layer(gis, settings)
    pending_ids = get_pending_submission_ids(submission_layer)

    if not pending_ids:
        LOGGER.info("No submissions are waiting to be processed.")
        return

    LOGGER.info("Found %d submission(s) to process.", len(pending_ids))
    for object_id in pending_ids:
        try:
            report = process_submission(submission_layer, object_id, settings)
            LOGGER.info("Created report for submission %s: %s", object_id, report)
        except Exception:
            # Continue so one invalid submission does not stop the batch.
            continue


if __name__ == "__main__":
    main()

